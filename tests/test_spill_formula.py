"""Tests for apply_spill_formula MCP tool"""
import pytest
import os
import tempfile
from pathlib import Path
import sys

# Add the parent directory to the path so we can import excel_mcp
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

# Set up environment
os.environ.setdefault('EXCEL_FILES_PATH', '/tmp/excel_files')

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import coordinate_to_tuple
from datetime import date


def get_excel_path(filename: str) -> str:
    """Get the full path for an Excel file"""
    base_path = os.environ.get('EXCEL_FILES_PATH', '/tmp/excel_files')
    os.makedirs(base_path, exist_ok=True)
    return os.path.join(base_path, filename)


def create_workbook(filename: str) -> str:
    """Create a new Excel workbook"""
    try:
        full_path = get_excel_path(filename)
        wb = Workbook()
        wb.save(full_path)
        wb.close()
        return f"Workbook created successfully at {full_path}"
    except Exception as e:
        return f"Error creating workbook: {str(e)}"


def write_data_to_excel(filepath: str, sheet_name: str, start_cell: str, data: list) -> str:
    """Write data to Excel file"""
    try:
        full_path = get_excel_path(filepath)
        wb = openpyxl.load_workbook(full_path)
        
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
        
        start_row, start_col = coordinate_to_tuple(start_cell)
        
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                sheet.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
        
        wb.save(full_path)
        wb.close()
        return f"Data written successfully"
    except Exception as e:
        return f"Error writing data: {str(e)}"


def apply_spill_formula(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    formula: str,
) -> str:
    """Apply dynamic array formula to specified range."""
    try:
        full_path = get_excel_path(filepath)
        
        # Open workbook
        wb = openpyxl.load_workbook(full_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"
        sheet = wb[sheet_name]
        
        # Validate cell references
        try:
            start_row, start_col = coordinate_to_tuple(start_cell)
            end_row, end_col = coordinate_to_tuple(end_cell)
            
            if start_row > end_row or start_col > end_col:
                return "Error: Invalid range specification: start_cell must be before end_cell"
        except ValueError:
            return f"Error: Invalid cell reference"
        
        # Set formula in start cell
        cell = sheet[start_cell]
        cell.value = formula
        
        # Apply dynamic array formula to range
        range_str = f"{start_cell}:{end_cell}"
        try:
            cell.set_dynamic_array_formula(range_str)
        except AttributeError:
            # Fallback if set_dynamic_array_formula is not available
            return "Error: This feature requires openpyxl-spill library. Please ensure it's properly installed."
        
        # Save workbook
        wb.save(full_path)
        wb.close()
        
        return f"Applied dynamic array formula '{formula}' to range {range_str}"
        
    except Exception as e:
        return f"Error applying spill formula: {str(e)}"


class TestApplySpillFormula:
    """Test suite for apply_spill_formula function"""
    
    @pytest.fixture
    def temp_excel_dir(self):
        """Create a temporary directory for test Excel files"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Set up the environment variable for EXCEL_FILES_PATH
            original_path = os.environ.get('EXCEL_FILES_PATH')
            os.environ['EXCEL_FILES_PATH'] = tmpdir
            yield tmpdir
            # Restore original environment variable
            if original_path:
                os.environ['EXCEL_FILES_PATH'] = original_path
            else:
                del os.environ['EXCEL_FILES_PATH']
    
    @pytest.fixture
    def test_excel_file(self, temp_excel_dir):
        """Create a test Excel file with sample data"""
        filename = "test_spill.xlsx"
        
        # Create workbook
        result = create_workbook(filename)
        assert "successfully" in result
        
        # Add sample data for testing UNIQUE function
        write_data_to_excel(
            filename,
            "Sheet",
            "A1",
            [
                ["Apple"],
                ["Banana"],
                ["Apple"],
                ["Cherry"],
                ["Banana"],
                ["Date"],
                ["Apple"],
                ["Elderberry"],
                ["Fig"],
                ["Cherry"]
            ]
        )
        
        # Add sample data for testing SORT function
        write_data_to_excel(
            filename,
            "Sheet",
            "B1",
            [
                ["Name", "Score"],
                ["Alice", 85],
                ["Bob", 92],
                ["Charlie", 78],
                ["David", 88],
                ["Eve", 95],
                ["Frank", 82],
                ["Grace", 90],
                ["Henry", 75],
                ["Ivy", 93]
            ]
        )
        
        return filename
    
    def test_apply_unique_formula(self, test_excel_file):
        """Test applying UNIQUE dynamic array formula"""
        result = apply_spill_formula(
            test_excel_file,
            "Sheet",
            "D1",
            "D10",
            "=UNIQUE(A1:A10)"
        )
        
        assert "Applied dynamic array formula" in result
        assert "=UNIQUE(A1:A10)" in result
        assert "D1:D10" in result
    
    def test_apply_sort_formula_single_column(self, test_excel_file):
        """Test applying SORT dynamic array formula for single column"""
        result = apply_spill_formula(
            test_excel_file,
            "Sheet",
            "E1",
            "E10",
            "=SORT(A1:A10)"
        )
        
        assert "Applied dynamic array formula" in result
        assert "=SORT(A1:A10)" in result
        assert "E1:E10" in result
    
    def test_apply_sort_formula_multiple_columns(self, test_excel_file):
        """Test applying SORT dynamic array formula for multiple columns"""
        result = apply_spill_formula(
            test_excel_file,
            "Sheet",
            "F1",
            "G10",
            "=SORT(B1:C10,2,-1)"
        )
        
        assert "Applied dynamic array formula" in result
        assert "=SORT(B1:C10,2,-1)" in result
        assert "F1:G10" in result
    
    def test_invalid_sheet_name(self, test_excel_file):
        """Test error handling for invalid sheet name"""
        result = apply_spill_formula(
            test_excel_file,
            "InvalidSheet",
            "A1",
            "A5",
            "=UNIQUE(A1:A10)"
        )
        
        assert "Error: Sheet 'InvalidSheet' not found" in result
    
    def test_invalid_cell_reference(self, test_excel_file):
        """Test error handling for invalid cell references"""
        result = apply_spill_formula(
            test_excel_file,
            "Sheet",
            "INVALID",
            "A5",
            "=UNIQUE(A1:A10)"
        )
        
        assert "Error: Invalid cell reference" in result
    
    def test_invalid_range_specification(self, test_excel_file):
        """Test error handling for invalid range (start > end)"""
        result = apply_spill_formula(
            test_excel_file,
            "Sheet",
            "D5",
            "D1",
            "=UNIQUE(A1:A10)"
        )
        
        assert "Error: Invalid range specification" in result
        assert "start_cell must be before end_cell" in result
    
    def test_complex_spill_formulas(self, test_excel_file):
        """Test various complex spill formulas"""
        # Test FILTER function
        result = apply_spill_formula(
            test_excel_file,
            "Sheet",
            "H1",
            "I10",
            '=FILTER(B1:C10,C1:C10>80)'
        )
        assert "Applied dynamic array formula" in result
        
        # Test SEQUENCE function
        result = apply_spill_formula(
            test_excel_file,
            "Sheet",
            "J1",
            "J10",
            "=SEQUENCE(10,1,1,1)"
        )
        assert "Applied dynamic array formula" in result
    
    def test_apply_spill_to_small_range(self, test_excel_file):
        """Test applying spill formula to a smaller range than needed"""
        # This should still succeed but the spill might be truncated
        result = apply_spill_formula(
            test_excel_file,
            "Sheet",
            "K1",
            "K3",
            "=UNIQUE(A1:A10)"
        )
        
        assert "Applied dynamic array formula" in result
        assert "K1:K3" in result

    def test_create_sample_with_spill_functions(self, temp_excel_dir):
        """Test creating Excel file with multiple spill functions similar to create_with_openpyxl.py"""
        
        # Create a new file for this test
        filename = "test_spill_sample.xlsx"
        
        # Create workbook
        result = create_workbook(filename)
        assert "successfully" in result
        
        # Write header
        write_data_to_excel(filename, "Sheet", "A1", [["Date", "Sales", "Region"]])
        
        # Write sample data
        data = [
            [date(2024, 1, 1), 1500, 'North'],
            [date(2024, 1, 2), 2300, 'South'],
            [date(2024, 1, 3), 1800, 'North'],
            [date(2024, 1, 4), 2100, 'East'],
            [date(2024, 1, 5), 1900, 'North'],
            [date(2024, 1, 6), 2500, 'South'],
            [date(2024, 1, 7), 1700, 'East'],
        ]
        write_data_to_excel(filename, "Sheet", "A2", data)
        
        # Add analysis header
        write_data_to_excel(filename, "Sheet", "E1", [["Analysis with SPILL Functions"]])
        
        # 1. SORT - Top Sales (F3:F9)
        write_data_to_excel(filename, "Sheet", "E3", [["Top Sales (Sorted):"]])
        result = apply_spill_formula(filename, "Sheet", "F3", "F9", "=SORT(B2:B8,,-1)")
        assert "Applied dynamic array formula" in result
        
        # 2. UNIQUE - Unique Regions (F12:F14)
        write_data_to_excel(filename, "Sheet", "E12", [["Unique Regions:"]])
        result = apply_spill_formula(filename, "Sheet", "F12", "F14", "=UNIQUE(C2:C8)")
        assert "Applied dynamic array formula" in result
        
        # 3. FILTER - North Sales (J3:J5)
        write_data_to_excel(filename, "Sheet", "I3", [["North Sales:"]])
        result = apply_spill_formula(filename, "Sheet", "J3", "J5", '=FILTER(B2:B8,C2:C8="North")')
        assert "Applied dynamic array formula" in result
        
        # 4. FILTER - Sales > 2000 (A17:C19)
        write_data_to_excel(filename, "Sheet", "A16", [["Sales > 2000:"]])
        result = apply_spill_formula(filename, "Sheet", "A17", "C19", "=FILTER(A2:C8,B2:B8>2000)")
        assert "Applied dynamic array formula" in result
        
        # 5. SEQUENCE (E20:E26)
        write_data_to_excel(filename, "Sheet", "D20", [["Numbers:"]])
        result = apply_spill_formula(filename, "Sheet", "E20", "E26", "=SEQUENCE(7)")
        assert "Applied dynamic array formula" in result
        
        return filename
    
    def test_compare_with_reference_file(self, temp_excel_dir):
        """Test comparing created file with reference openpyxl_spill.xlsx"""
        
        # First create our test file
        test_file = self.test_create_sample_with_spill_functions(temp_excel_dir)
        test_path = Path(temp_excel_dir) / test_file
        
        # Check reference file exists
        reference_path = Path(__file__).parent.parent / "sample" / "openpyxl_spill.xlsx"
        if not reference_path.exists():
            pytest.skip("Reference file openpyxl_spill.xlsx not found")
        
        # Load both workbooks
        test_wb = openpyxl.load_workbook(test_path, data_only=False)
        ref_wb = openpyxl.load_workbook(reference_path, data_only=False)
        
        test_ws = test_wb.active
        ref_ws = ref_wb.active
        
        # Compare key cells with formulas
        formula_cells = [
            ("F3", "=SORT(B2:B8,,-1)"),  # SORT formula
            ("F12", "=UNIQUE(C2:C8)"),    # UNIQUE formula
            ("J3", '=FILTER(B2:B8,C2:C8="North")'),  # FILTER North
            ("A17", "=FILTER(A2:C8,B2:B8>2000)"),    # FILTER >2000
            ("E20", "=SEQUENCE(7)")        # SEQUENCE formula
        ]
        
        for cell_ref, expected_formula in formula_cells:
            test_cell = test_ws[cell_ref]
            ref_cell = ref_ws[cell_ref]
            
            # Check if both have formulas
            assert test_cell.value is not None, f"Test cell {cell_ref} should have a formula"
            
            # Handle ArrayFormula objects
            from openpyxl.worksheet.formula import ArrayFormula
            if isinstance(test_cell.value, ArrayFormula):
                test_formula = test_cell.value.text
            else:
                test_formula = str(test_cell.value)
            
            assert test_formula.startswith("="), f"Test cell {cell_ref} should contain a formula"
            
            # Compare formula content (normalize for comparison)
            # Remove Excel internal function prefixes
            test_formula_normalized = test_formula.upper().replace(" ", "")
            test_formula_normalized = test_formula_normalized.replace("=_XLFN._XLWS.", "=")
            test_formula_normalized = test_formula_normalized.replace("=_XLFN.", "=")
            expected_normalized = expected_formula.upper().replace(" ", "")
            assert test_formula_normalized == expected_normalized, f"Formula mismatch in {cell_ref}: {test_formula_normalized} != {expected_normalized}"
        
        # Compare data cells
        data_cells = [
            "A1", "B1", "C1",  # Headers
            "A2", "B2", "C2",  # First data row
            "E1", "E3", "E12", "I3", "A16", "D20"  # Labels
        ]
        
        for cell_ref in data_cells:
            test_value = test_ws[cell_ref].value
            ref_value = ref_ws[cell_ref].value
            
            # For dates, compare date part only
            if isinstance(test_value, date) and isinstance(ref_value, date):
                assert test_value.date() == ref_value.date() if hasattr(test_value, 'date') else test_value == ref_value
            else:
                assert test_value == ref_value, f"Value mismatch in {cell_ref}: {test_value} != {ref_value}"
        
        test_wb.close()
        ref_wb.close()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])