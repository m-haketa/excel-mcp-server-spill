#!/usr/bin/env python3
"""create_with_openpyxl.py - openpyxlでスピル関数ファイルを作成"""

import sys
import os

# ローカルの修正済みopenpyxlを使用（必要に応じてパスを調整）
# sys.path.insert(0, os.path.abspath('/workspace/openpyxl'))

from openpyxl import Workbook
from datetime import date

print("Creating Excel file with SPILL functions using openpyxl...")

# ワークブックを作成
wb = Workbook()
ws = wb.active

# サンプルデータ
ws['A1'] = 'Date'
ws['B1'] = 'Sales'
ws['C1'] = 'Region'

data = [
    (date(2024, 1, 1), 1500, 'North'),
    (date(2024, 1, 2), 2300, 'South'),
    (date(2024, 1, 3), 1800, 'North'),
    (date(2024, 1, 4), 2100, 'East'),
    (date(2024, 1, 5), 1900, 'North'),
    (date(2024, 1, 6), 2500, 'South'),
    (date(2024, 1, 7), 1700, 'East'),
]

# データを書き込み
for row, (dt, sales, region) in enumerate(data, start=2):
    ws[f'A{row}'] = dt
    ws[f'B{row}'] = sales
    ws[f'C{row}'] = region

# SPILL関数の追加（範囲が重ならないように配置）
print("\nAdding SPILL functions with non-overlapping ranges...")

# 分析セクションのヘッダー
ws['E1'] = 'Analysis with SPILL Functions'

# 1. SORT - 列F（F3から下へスピル）
ws['E3'] = 'Top Sales (Sorted):'
cell = ws['F3']
cell.value = '=SORT(B2:B8,,-1)'
cell.set_dynamic_array_formula('F3:F9')
print("✓ SORT formula added (F3:F9)")

# 2. UNIQUE - 列H（H3から下へスピル）
ws['E12'] = 'Unique Regions:'
cell = ws['F12']
cell.value = '=UNIQUE(C2:C8)'
cell.set_dynamic_array_formula('F12:F14')
print("✓ UNIQUE formula added (F12:F14)")

# 3. FILTER (North) - 列J（J3から下へスピル）
ws['I3'] = 'North Sales:'
cell = ws['J3']
cell.value = '=FILTER(B2:B8,C2:C8="North")'
cell.set_dynamic_array_formula('J3:J5')
print("✓ FILTER formula added (J3:J5)")

# 4. FILTER (>2000) - 複数列、行17から（A17:C19）
ws['A16'] = 'Sales > 2000:'
cell = ws['A17']
cell.value = '=FILTER(A2:C8,B2:B8>2000)'
cell.set_dynamic_array_formula('A17:C19')
print("✓ FILTER formula added (A17:C19) - multi-column")

# 5. SEQUENCE - 列E、行20から（E20:E26）
ws['D20'] = 'Numbers:'
cell = ws['E20']
cell.value = '=SEQUENCE(7)'
cell.set_dynamic_array_formula('E20:E26')
print("✓ SEQUENCE formula added (E20:E26)")

# ワークブックを保存
output_path = os.path.join(os.path.dirname(__file__), 'openpyxl_spill_generated.xlsx')
wb.save(output_path)

print(f"\n✓ openpyxl file created: {output_path}")